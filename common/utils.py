import json
import os
from openpyxl import load_workbook


def get_path():
    """
    获取当前路径的父路径
    :return: dir
    """
    current_dir = os.path.abspath(__file__)
    parent_dir = os.path.dirname(os.path.dirname(current_dir))
    json_path = parent_dir + r'\data\topic_data.json'
    excel_path = parent_dir + r'\data\topic_data.xlsx'
    log_path = parent_dir + r'\logs\log.txt'
    return json_path, excel_path, log_path


def get_json_data(params):
    """
    根据测试方法名，读取json格式参数化参数
    :param params: 测试方法名
    :return: 参数化数据
    """
    with open(get_path()[0], 'r', encoding='utf8') as file:
        data = file.read()
        topic_basic_info = json.loads(data)
        return topic_basic_info[params]


def get_csv_data():
    """
    读取Excel格式参数化数据
    :return: 参数化数据
    """
    wb = load_workbook(get_path()[1])
    # print(wb.worksheets)
    # 获取data页签的数据
    ws = wb['data']
    testcase_data = []
    # 对行进行遍历即遍历2 - 4行
    # 获取2 - 4行 第4列的数据
    # 注意这里不是索引，而是行数和列数，都是从1开始
    for x in range(2, len(tuple(ws.rows)) + 1):
        """
        ws.cell(row=1, column=1).value 可以用来读取第一行第一列单元格的内容，
        ws.cell(row=1, column=1).value = 'new value' 可以用来修改这个单元格的内容
        """
        # 反序列化
        get_data = json.loads(ws.cell(row=x, column=4).value)
        testcase_data.append(get_data)

    return testcase_data


if __name__ == '__main__':
    print(get_csv_data())